import pandas as pd
import datetime
from flask import Flask, render_template, request, redirect, url_for
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from tabulate import tabulate

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Verifica se o arquivo foi enviado
        if 'file' not in request.files:
            return "Nenhum arquivo enviado"
        
        file = request.files['file']
        
        if file.filename == '':
            return "Nome de arquivo inválido"
        
        if file and allowed_file(file.filename):
            filename = file.filename
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Processa o arquivo
            processar_arquivo(filepath)
            
            return redirect(url_for('resultado', filename=filename))
    
    return render_template('index.html')

@app.route('/resultado/<filename>')
def resultado(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    # Carrega o arquivo processado
    df = pd.read_excel(filepath)
    
    # Formata a tabela
    tabela = tabulate(df[['Item', 'Operador', 'Termino', 'Tempo restante', 'Tempo excedente']], 
                     headers='keys', tablefmt='html')
    
    return f"""
    <h1>Resultados:</h1>
    {tabela}
    <br>
    <a href="/">Voltar</a>
    """

def processar_arquivo(arquivo_excel):
    """Processa o arquivo Excel"""
    df = pd.read_excel(arquivo_excel)

    for index, row in df.iterrows():
        if pd.notna(row['Termino']):
            termino_str = str(row['Termino'])
            try:
                termino = datetime.datetime.strptime(termino_str, '%d/%m/%Y %H:%M:%S')
            except ValueError:
                try:
                    hora = datetime.datetime.strptime(termino_str, '%H:%M:%S').time()
                    data_atual = datetime.date.today()
                    termino = datetime.datetime.combine(data_atual, hora)
                except ValueError:
                    print(f"Formato inválido na linha {index + 2}: {termino_str}")
                    continue

            agora = datetime.datetime.now()

            if termino > agora:
                tempo_restante = termino - agora
                df.loc[index, 'Tempo restante'] = str(tempo_restante)
                df.loc[index, 'Tempo excedente'] = ''
            else:
                tempo_excedente = agora - termino
                df.loc[index, 'Tempo restante'] = ''
                df.loc[index, 'Tempo excedente'] = str(tempo_excedente)
        else:
            df.loc[index, 'Tempo restante'] = ''
            df.loc[index, 'Tempo excedente'] = ''

    # Salva o arquivo processado
    df.to_excel(arquivo_excel, index=False)
    
    # Aplica formatação condicional
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        rule = CellIsRule(operator="equal", formula=['0'], fill=red_fill)
        ws.conditional_formatting.add(f"D2:D{len(df) + 1}", rule)
        wb.save(arquivo_excel)
    except Exception as e:
        print("Erro na formatação:", str(e))

if __name__ == "__main__":
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True)